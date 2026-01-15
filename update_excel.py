import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

STOCKS = {
    "2330.TW": "2330",
    "0050.TW": "0050",
    "006208.TW": "006208"
}

START_DATE = "2020-01-02"
END_DATE = datetime.today().strftime("%Y-%m-%d")
FILE_PATH = "docs/台股每日紀錄.xlsx"

writer = pd.ExcelWriter(FILE_PATH, engine="openpyxl")

for ticker, sheet in STOCKS.items():
    df = yf.download(ticker, start=START_DATE, end=END_DATE)
    df = df[["Open", "High", "Low", "Close"]]
    df.reset_index(inplace=True)

    df.columns = ["日期", "開盤價", "最高價", "最低價", "收盤價"]
    df["漲跌幅%"] = df["收盤價"].pct_change() * 100
    df["漲跌幅%"] = df["漲跌幅%"].round(2)

    df.to_excel(writer, sheet_name=sheet, index=False)

writer.close()

wb = load_workbook(FILE_PATH)

for sheet in STOCKS.values():
    ws = wb[sheet]
    ws["H1"] = "區間選擇"
    ws["H2"] = "近一個月"
    ws["H4"] = "區間漲跌幅%"

    dv = DataValidation(
        type="list",
        formula1='"近一個月,近三個月,近半年,近一年"',
        allow_blank=False
    )
    ws.add_data_validation(dv)
    dv.add("H2")

    last = ws.max_row
    ws["H5"] = (
        f'=IF(H2="近一個月",(E{last}/E{last-22}-1)*100,'
        f'IF(H2="近三個月",(E{last}/E{last-66}-1)*100,'
        f'IF(H2="近半年",(E{last}/E{last-132}-1)*100,'
        f'(E{last}/E{last-252}-1)*100)))'
    )

wb.save(FILE_PATH)
